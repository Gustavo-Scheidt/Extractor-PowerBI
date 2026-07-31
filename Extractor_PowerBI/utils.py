import glob
import logging
import os
import re
import shutil

import pandas as pd
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from config import (
    ARQUIVO_LOG,
    PASTA_CSV,
    PASTA_EXCEL,
    PASTA_LOGS,
    PASTA_PAYLOADS,
    PASTA_RESPOSTAS,
)


def configurar_logging():
    """Configura o logger raiz do projeto: log em arquivo + console."""
    os.makedirs(PASTA_LOGS, exist_ok=True)

    logger = logging.getLogger("extractor_powerbi")
    logger.setLevel(logging.INFO)

    if logger.handlers:
        return logger  # já configurado (evita handlers duplicados)

    formato = logging.Formatter("[%(asctime)s] %(levelname)s: %(message)s", "%d/%m/%Y %H:%M:%S")

    file_handler = logging.FileHandler(ARQUIVO_LOG, encoding="utf-8")
    file_handler.setFormatter(formato)

    console_handler = logging.StreamHandler()
    console_handler.setFormatter(formato)

    logger.addHandler(file_handler)
    logger.addHandler(console_handler)

    return logger


def criar_pastas():
    """Cria todas as pastas de saída necessárias para o funcionamento do extractor."""
    pastas = [PASTA_PAYLOADS, PASTA_RESPOSTAS, PASTA_CSV, PASTA_EXCEL, PASTA_LOGS]
    for pasta in pastas:
        os.makedirs(pasta, exist_ok=True)


def limpar_pastas_antigas():
    """Limpa os arquivos temporários das execuções anteriores."""
    logger = logging.getLogger("extractor_powerbi")
    pastas = [PASTA_PAYLOADS, PASTA_RESPOSTAS, PASTA_CSV]
    for pasta in pastas:
        if os.path.exists(pasta):
            for f in os.listdir(pasta):
                caminho = os.path.join(pasta, f)
                try:
                    if os.path.isfile(caminho) or os.path.islink(caminho):
                        os.remove(caminho)
                    elif os.path.isdir(caminho):
                        shutil.rmtree(caminho, ignore_errors=True)
                except Exception as e:
                    logger.warning("Aviso ao limpar %s: %s", caminho, e)


def listar_payloads():
    """Lista todos os arquivos JSON de payload capturados na pasta."""
    padrao = os.path.join(PASTA_PAYLOADS, "payload_*.json")
    return sorted(glob.glob(padrao))


def _nome_legivel(nome_payload, df, usados, limite=31, caminho_payload=None):
    """
    Gera um nome legível a partir dos nomes reais das colunas do DataFrame
    (ex.: 'Polo_VBP_ano_base_2021'). Cai para o nome do payload quando as
    colunas ainda são genéricas (0, 1, 2...). Garante nomes únicos mesmo
    após o corte de tamanho (relevante para abas do Excel, limitadas a 31
    caracteres).
    """
    colunas_nomeadas = [str(c) for c in df.columns if not str(c).isdigit()]
    base = "_".join(colunas_nomeadas) if colunas_nomeadas else nome_payload
    base = re.sub(r'[\\/*?:\[\]<>|]', '_', base).strip("_")
    base = re.sub(r'_+', '_', base) or nome_payload

    nome = base[:limite]
    contador = 1
    nome_final = nome
    while nome_final in usados:
        sufixo = f"_{contador}"
        nome_final = nome[: limite - len(sufixo)] + sufixo
        contador += 1

    usados.add(nome_final)
    return nome_final


def _formatar_planilha(worksheet, df):
    """Cabeçalho em negrito, colunas com largura ajustada e primeira linha fixa."""
    worksheet.freeze_panes = "A2"

    for col_idx, col in enumerate(df.columns, start=1):
        worksheet.cell(row=1, column=col_idx).font = Font(bold=True)

        # CORREÇÃO AQUI: uso de .apply(lambda x: len(str(x))) evita o erro TypeError no Pandas
        maior_valor = df[col].apply(lambda x: len(str(x))).max() if not df.empty else 0
        
        largura = max(len(str(col)), maior_valor) + 2
        worksheet.column_dimensions[get_column_letter(col_idx)].width = min(largura, 50)


def salvar_excel_consolidado(dataframes_dict, caminho_excel):
    """
    Salva todas as abas no arquivo Excel, com nomes legíveis, cabeçalho
    formatado, colunas ajustadas e uma aba 'Resumo' listando payload → aba
    → linhas/colunas de cada visual.
    """
    os.makedirs(os.path.dirname(caminho_excel), exist_ok=True)

    usados = set()
    resumo = []

    with pd.ExcelWriter(caminho_excel, engine="openpyxl") as writer:
        for nome_payload, df in dataframes_dict.items():
            if df is None or df.empty:
                continue

            nome_aba = _nome_legivel(nome_payload, df, usados, limite=31)
            df.to_excel(writer, sheet_name=nome_aba, index=False)
            _formatar_planilha(writer.sheets[nome_aba], df)

            resumo.append(
                {
                    "Payload": nome_payload,
                    "Aba": nome_aba,
                    "Linhas": df.shape[0],
                    "Colunas": df.shape[1],
                }
            )

        if resumo:
            df_resumo = pd.DataFrame(resumo)
            df_resumo.to_excel(writer, sheet_name="Resumo", index=False)
            _formatar_planilha(writer.sheets["Resumo"], df_resumo)
            # Move "Resumo" para ser a primeira aba do arquivo
            writer.book.move_sheet("Resumo", offset=-(len(writer.book.sheetnames) - 1))


def salvar_csvs_individuais(dataframes_dict, pasta_csv):
    """Salva cada visual como um CSV individual em pasta_csv/<nome_legivel>.csv."""
    os.makedirs(pasta_csv, exist_ok=True)
    usados = set()
    caminhos = []

    for nome_payload, df in dataframes_dict.items():
        if df is not None and not df.empty:
            nome_arquivo = _nome_legivel(nome_payload, df, usados, limite=80)
            caminho = os.path.join(pasta_csv, f"{nome_arquivo}.csv")
            df.to_csv(caminho, index=False, encoding="utf-8-sig")
            caminhos.append(caminho)

    return caminhos